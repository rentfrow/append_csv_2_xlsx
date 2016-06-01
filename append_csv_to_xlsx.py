#!/usr/bin/env python3

""" 
    Script appends csv file contents to an existing Excel xlsx file

    Supports Excel 2013 xlsx format
    CSV file format must be comma separated

    Please note!!! 
        This script will strip any existing charts and photos from 
        the target xlsx file.
    
   ToDo:
       - Enable importing multiple csv files at one time. - done
       - Enable multiple csv formats (Comma, Tab, Space, ...)
       - Create a template Summary page for created xlsx files
       - Add time in minutes option - partially done, currently mandatory
       - Add normalize time option
       - Search for a template or summary directory and use that to create a summary page
       - Search for a picture directory and import pictures into xlsx workbook
       - milliseconds not working
       - some csv files have a non csv header with file information. Currently the script 
         bombs on this. Enable it detect the csv file header contents and adjust. 
 
    Author: Brad Rentfrow
            brentfro@cisco.com
            brad@rentfrow.us
    Last Update: Jun01-2016
"""

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Color, colors, PatternFill
from openpyxl.cell import get_column_letter
import csv
import argparse
import os
import sys
import datetime
import re

# need to make this a command line option
elapsed_time_opt = True

# Collect command arguments and parse into variables
def collect_args():
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx_append_to_file", help="output xlsx file to append to")
    parser.add_argument("csv_input_file", nargs='*', help="csv file to import")
    args = parser.parse_args()
    return args
    
# Open target xlsx file
def open_xlsx_file(file_dest):
    print("Opening: " + file_dest)
    wb = load_workbook(filename = file_dest)
    return wb
    
# Check destination xlsx file and create if does not exist
def check_dest_xlxs(filename):
    if os.path.isfile(filename):
        print("Target file " + filename + " exists")
    else:
        print("Target file " + filename + " does not exist... creating.")
        # Create a workbook with a single worksheet call Summary
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Summary"
        header_font = Font(color=colors.RED, size=24, bold=True)
        yellow_fill = PatternFill(start_color=colors.YELLOW,
                   end_color=colors.YELLOW,
                   fill_type='solid')
        ws1['A1'].font = header_font
        ws1['A1'].fill = yellow_fill
        ws1['A1'] = filename + " Summary"
        # Set the column sizes
        ws1.column_dimensions['A'].width = 50
        ws1.column_dimensions['B'].width = 50        
        wb.save(filename)

# Read in CSV file and append contents to new worksheet on specified workbook
def append_csv_to_xlsx(cvs_src, wb):
    sheet = strip_ext(cvs_src)
    new_ws = wb.create_sheet(title=sheet)
    print("Reading: " + cvs_src)
    with open(cvs_src) as csv_in:
        #read the csv
        reader = csv.reader(csv_in)
        if elapsed_time_opt:
            #enumerate the rows, so that you can
            #get the row index for the xlsx
            for row_index, row in enumerate(reader):
                # enumerate the row list so that you can
                # get the column index for the xlsx
                for col_index, cel_value in enumerate(row):
                    cel_col = col_index + 1
                    cel_row = row_index + 1
                    # Print the first column 
                    if cel_col == 1:
                        if cel_row == 1:
                            try:
                                new_ws.cell(column=cel_col, row=cel_row).value = float(cel_value)
                            except ValueError:
                                new_ws.cell(column=cel_col, row=cel_row).value = cel_value
                        if cel_row == 2:
                            myd = csv_cnvt_date(cel_value)
                            #print(myd[0], myd[1], myd[2], myd[3], myd[4], myd[5], myd[6])
                            new_ws.cell(column=cel_col, row=cel_row).value = datetime.datetime(myd[0], myd[1], myd[2], myd[3], myd[4], myd[5], myd[6])
                            new_ws.cell(column=cel_col, row=cel_row + 1).value = datetime.datetime(myd[0], myd[1], myd[2], myd[3], myd[4], myd[5], myd[6])
                        if cel_row >= 3:
                            myd = csv_cnvt_date(cel_value)
                            #print(myd[0], myd[1], myd[2], myd[3], myd[4], myd[5], myd[6])
                            #new_ws.cell(column=cel_col, row=cel_row).value = datetime.datetime(myd[0], myd[1], myd[2], myd[3], myd[4], myd[5], myd[6])
                            new_ws.cell(column=cel_col, row=cel_row + 1).value = datetime.datetime(myd[0], myd[1], myd[2], myd[3], myd[4], myd[5], myd[6])             
                    # Print insert an elapsed time column and the 2 column data on the third column
                    elif cel_col == 2:
                        if cel_row == 1:
                            # Inserting elasped time column here - header goes first
                            new_ws.cell(column=cel_col, row=cel_row).value = "time(min)"
                            # put column two header data on column three
                            new_ws.cell(column=cel_col + 1, row=cel_row).value = cel_value
                        if cel_row == 2:
                            # Insert 0 to begin elapsed time
                            new_ws.cell(column=cel_col, row=cel_row).value = float(0)
                            new_ws.cell(column=cel_col, row=cel_row).number_format = '0.00'
                            # put column two data on column three
                            try:
                               new_ws.cell(column=cel_col + 1, row=cel_row).value = float(cel_value)
                               new_ws.cell(column=cel_col + 1, row=cel_row).number_format = '0.00'
                               new_ws.cell(column=cel_col + 1, row=cel_row + 1).value = float(cel_value)
                               new_ws.cell(column=cel_col + 1, row=cel_row + 1).number_format = '0.00'
                            except ValueError:
                               new_ws.cell(column=cel_col + 1, row=cel_row).value = cel_value
                               new_ws.cell(column=cel_col + 1, row=cel_row + 1).value = cel_value
                        if cel_row >= 3:
                            # Building formula "= 60 * 24 * ( A3 - A2 ) + B2"
                            time_1 = get_column_letter(cel_col - 1) + str(cel_row)
                            time_2 = get_column_letter(cel_col - 1) + str(cel_row - 1)
                            e_time = get_column_letter(cel_col) + str(cel_row - 1)
                            e_time_fomula = "=60*24*(" + time_1 + "-" + time_2 + ")+" + e_time
                            # Insert formula starting on 'B3' or cel_col 2, cel_row 3
                            new_ws.cell(column=cel_col, row=cel_row).value = e_time_fomula
                            new_ws.cell(column=cel_col, row=cel_row).number_format = '0.00'
                            # put column two data on column three again... 
                            try:
                                new_ws.cell(column=cel_col + 1, row=cel_row + 1).value = float(cel_value)
                                new_ws.cell(column=cel_col + 1, row=cel_row + 1).number_format = '0.00'
                            except ValueError:
                                new_ws.cell(column=cel_col + 1, row=cel_row + 1).value = cel_value
                    # Print the rest of the columns
                    elif cel_col >= 3:
                        cel_col = cel_col + 1
                        if cel_row == 1:
                            # Print the header row
                            new_ws.cell(column=cel_col, row=cel_row).value = cel_value
                        if cel_row == 2:
                            try:
                                new_ws.cell(column=cel_col, row=cel_row).value = float(cel_value)
                                new_ws.cell(column=cel_col, row=cel_row).number_format = '0.00'
                                new_ws.cell(column=cel_col, row=cel_row + 1).value = float(cel_value)
                                new_ws.cell(column=cel_col, row=cel_row + 1).number_format = '0.00'
                            except ValueError:
                                new_ws.cell(column=cel_col, row=cel_row).value = cel_value
                                new_ws.cell(column=cel_col, row=cel_row + 1).value = cel_value
                            #cel_row = cel_row + 1
                            try:
                                new_ws.cell(column=cel_col, row=cel_row + 1).value = float(cel_value)
                                new_ws.cell(column=cel_col, row=cel_row + 1).number_format = '0.00'
                            except ValueError:
                                new_ws.cell(column=cel_col, row=cel_row + 1).value = cel_value
                        if cel_row >= 3:
                            try:
                                new_ws.cell(column=cel_col, row=cel_row + 1).value = float(cel_value)
                                new_ws.cell(column=cel_col, row=cel_row + 1).number_format = '0.00'
                            except ValueError:
                                new_ws.cell(column=cel_col, row=cel_row + 1).value = cel_value
                    # Convert value to float or just print as is
                    # try:
                        # new_ws.cell(column=cel_col, row=cel_row).value = float(cel_value)
                    # except ValueError:
                        # new_ws.cell(column=cel_col, row=cel_row).value = cel_value
        else:
            #enumerate the rows, so that you can
            #get the row index for the xlsx
            for index, row in enumerate(reader):
                # enumerate the row list so that you can
                # get the column index for the xlsx
                for col, cel_value in enumerate(row):
                    cel_col = col + 1
                    cel_row = index + 1
                    if cel_col == 1:
                        if cel_row != 1:
                            myd = csv_cnvt_date(cel_value)
                            # print(myd[0], myd[1], myd[2], myd[3], myd[4], myd[5], myd[6])
                            new_ws.cell(column=cel_col, row=cel_row).value = datetime.datetime(myd[0], myd[1], myd[2], myd[3], myd[4], myd[5], myd[6])
                    else:
                        # Convert value to float or just print as is
                        try:
                            new_ws.cell(column=cel_col, row=cel_row).value = float(cel_value)
                        except ValueError:
                            new_ws.cell(column=cel_col, row=cel_row).value = cel_value

# Month to digit
def month_to_digit(month):
    monthdigit = {'Jan': '01',
                    'Feb': '02',
                    'Mar': '03',
                    'Apr': '04',
                    'May': '05',
                    'Jun': '06',
                    'Jul': '07',
                    'Aug': '08',
                    'Sep': '09',
                    'Oct': '10',
                    'Nov': '11',
                    'Dec': '12'}
    return monthdigit[month]
    
# Fix Date
# look into replacing with strptime module
def csv_cnvt_date(the_datestamp):
    # print("Processing date: " + the_datestamp)
    # Apr 07, 2016 21:56:43.043
    if re.match(r'[JanFebMarApyulgSOctNovDc]{3}\s\d{2},\s\d{4}\s\d{2}:\d{2}:\d{2}\.\d{3}', the_datestamp, re.M) is not None:
        the_date_time = the_datestamp.split( )
        the_year = the_date_time[2]
        the_month = month_to_digit(the_date_time[0])
        the_date = the_date_time[1].split(',')
        the_day = the_date[0]
        the_time = the_date_time[3].split(':')
        the_hours = the_time[0]
        the_minutes = the_time[1]
        the_seconds = the_time[2].split('.')
        the_sec = int(the_seconds[0]) 
        the_msec = the_seconds[1]
        # the_msec = float(the_seconds[1]) / 1000 # still not working just prints zero
    # date = 04/29/16 21:13:41.002
    elif re.match(r'^\d{1,2}/\d{1,2}/\d{2,2}\s\d{1,2}:\d{1,2}:\d{1,2}\.\d{1,3}', the_datestamp, re.M) is not None:
        the_date_time = the_datestamp.split( )
        the_date = the_date_time[0].split('/')
        the_time = the_date_time[1].split(':')
        the_date[2] = "20" + str(the_date[2]) # adding 20 suffix to year as datetime requires it
        the_year = the_date[2]
        the_month = the_date[0]
        the_day = the_date[1]
        the_hours = the_time[0]
        the_minutes = the_time[1]
        the_seconds = the_time[2].split('.')
        the_sec = int(the_seconds[0]) 
        # the_msec = int(the_seconds[1])
        the_msec = float(the_seconds[1]) / 1000 # still not working just prints zero
    # date = 04/29/16 21:13:41
    elif re.match(r'^\d{1,2}/\d{1,2}/\d{2,2}\s\d{1,2}:\d{1,2}:\d{1,2}$', the_datestamp, re.M) is not None:
        the_date_time = the_datestamp.split( )
        the_date = the_date_time[0].split('/')
        the_time = the_date_time[1].split(':')
        the_date[2] = "20" + str(the_date[2]) # adding 20 suffix to year as datetime requires it
        the_year = the_date[2]
        the_month = the_date[0]
        the_day = the_date[1]
        the_hours = the_time[0]
        the_minutes = the_time[1]
        the_sec = int(the_time[2]) 
        the_msec = 0
    # date = 04/29/2016 21:13:41
    elif re.match(r'^\d{1,2}/\d{1,2}/\d{4}\s\d{1,2}:\d{1,2}:\d{1,2}$', the_datestamp, re.M) is not None:
        the_date_time = the_datestamp.split( )
        the_date = the_date_time[0].split('/')
        the_time = the_date_time[1].split(':')
        the_year = the_date[2]
        the_month = the_date[0]
        the_day = the_date[1]
        the_hours = the_time[0]
        the_minutes = the_time[1]
        the_sec = int(the_time[2]) 
        the_msec = 0
    # date = 4/5/2016 21:27
    elif re.match(r'^\d{1,2}/\d{1,2}/\d{2,4}\s\d{1,2}:\d{1,2}$', the_datestamp, re.M) is not None:
        the_date_time = the_datestamp.split( )
        the_date = the_date_time[0].split('/')
        the_time = the_date_time[1].split(':')
        the_year = the_date[2]
        the_month = the_date[0]
        the_day = the_date[1]
        the_hours = the_time[0]
        the_minutes = the_time[1]
        the_sec = 0
        the_msec = 0
    # No timestamp detected
    else:
        the_year = 1900
        the_month = 1
        the_day = 1
        the_hours = 0
        the_minutes = 0
        the_sec = 0
        the_msec = 0
        #print(int(the_date[2]) , int(the_date[0]), int(the_date[1]), int(the_time[0]), int(the_time[1]), the_sec, the_msec)
    return [
            int(the_year),
            int(the_month),
            int(the_day),
            int(the_hours),
            int(the_minutes),
            int(the_sec), 
            int(the_msec) 
            ]
    
# Strips the file name extension
def strip_ext(file_name):
    fn = file_name.split('.')
    return fn[0]

# Save the XLSX file with the appened csv file contents
def save_workbook(wb, xlsx_file_name):
    print("Saving: " + xlsx_file_name)
    wb.save(xlsx_file_name)

# check incoming csv file name if it has a * adjust for a wild card
def check_incoming_csv(file, wb):
    for name in file:
        append_csv_to_xlsx(name, wb)
        print("Appending: %s" % name)
    
def check_elasped_time_opt(eto):
    if eto:
        print(eto)
    return eto
        
def main():
    print("### Appending csv files to xlsx file ###")
    args = collect_args()
    check_dest_xlxs(args.xlsx_append_to_file)
    wb = open_xlsx_file(args.xlsx_append_to_file)
    check_incoming_csv(args.csv_input_file, wb)
    ## append_csv_to_xlsx(args.csv_input_file, wb)
    save_workbook(wb, args.xlsx_append_to_file)
    print("### Done ###")

main()

